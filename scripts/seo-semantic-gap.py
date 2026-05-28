#!/usr/bin/env python3
"""
Generate the semantic-gap audit CSV from a Semrush "Ideas" export.

Companion to scripts/seo-keyword-audit.mjs (page-keyword coverage) and
scripts/seo-backlinks-outreach.py (off-site outreach). Same source,
different slice: the "related words are not present" idea family.

Semrush's semantic-gap idea ships as a sentence with a list of
backtick-wrapped phrases — the related words it judges your competitors
are using and you aren't. This script extracts those phrases, dedupes
per (url, keyword), and emits a per-page CSV the merchant can use to
know which topics to weave into the page content.

Output columns:
  url                       — site-relative path
  target_keyword            — the keyword the page is ranking for
  related_words_missing     — comma-joined list of phrases Semrush
                              wants you to add. Each phrase is a
                              concrete semantic concept; covering ~3-5
                              of them per page is usually enough to
                              close the gap.
  current_seo_title         — what's there now (from data/url-inventory/)
  suggested_action          — short heuristic: "FAQ section" /
                              "intro paragraph" / "subsection" based
                              on how many phrases are missing.

Usage:
  python3 scripts/seo-semantic-gap.py path/to/semrush-export.xlsx \\
    > docs/seo-audits/semantic-gap-YYYYMMDD.csv

Stderr prints a one-line per-tier summary.

Why surface this vs. auto-edit the bodies: weaving 5-10 new phrases
into existing prose without changing the merchant's voice requires
content judgment. The CSV makes the merchant's edit fast (they see
exactly which phrases to add and the current title for context) and
keeps voice control with the merchant.
"""

import csv
import json
import re
import sys
from collections import defaultdict
from openpyxl import load_workbook


def load_inventory():
    """Pull current seo.title from data/url-inventory/. Empty when missing."""
    out = {}
    for key, path in (('collections', 'data/url-inventory/collections.json'),
                      ('pages', 'data/url-inventory/pages.json')):
        try:
            with open(path) as f:
                data = json.load(f)
            for c in data.get(key, []):
                out[f'/{key}/{c["handle"]}'] = c.get('seoTitle') or c.get('title') or ''
        except Exception as e:
            sys.stderr.write(f'[warn] {path} not loaded: {e}\n')
    try:
        with open('data/url-inventory/blogs.json') as f:
            data = json.load(f)
        for b in data.get('blogs', []):
            for a in b.get('articles', []) or []:
                out[f'/blogs/{b["handle"]}/{a["handle"]}'] = a.get('title') or ''
    except Exception as e:
        sys.stderr.write(f'[warn] blogs.json not loaded: {e}\n')
    return out


def main():
    if len(sys.argv) < 2:
        sys.stderr.write('Usage: seo-semantic-gap.py <semrush-ideas.xlsx>\n')
        sys.exit(2)

    wb = load_workbook(sys.argv[1], data_only=True)
    ws = wb['SEO Ideas'] if 'SEO Ideas' in wb.sheetnames else wb[wb.sheetnames[0]]

    site_prefixes = (
        'https://www.mattressstoreslosangeles.com',
        'https://mattressstoreslosangeles.com',
    )

    # Dedupe per (url, keyword) — Semrush can flag the same combo
    # across multiple rows when content has multiple gap clusters.
    by_key = defaultdict(set)
    for r in range(2, ws.max_row + 1):
        idea = ws.cell(row=r, column=4).value or ''
        url = (ws.cell(row=r, column=2).value or '')
        kw = ws.cell(row=r, column=3).value or ''
        if 'related words are not present' not in idea:
            continue
        for prefix in site_prefixes:
            url = url.replace(prefix, '')
        # Extract everything inside `backticks` from the idea text.
        phrases = re.findall(r'`([^`]+)`', idea)
        if not phrases:
            continue
        by_key[(url, kw)].update(p.strip() for p in phrases if p.strip())

    inv = load_inventory()

    rows = []
    for (url, kw), phrases in by_key.items():
        title = inv.get(url, '')
        # Heuristic: more than 5 missing phrases → suggest a FAQ section
        # (each phrase fits naturally as a question topic). 3-5 → an
        # intro paragraph rewrite. 1-2 → an inline sentence addition.
        n = len(phrases)
        action = 'FAQ section (5+ Q&As)' if n >= 5 else (
            'Intro paragraph rewrite' if n >= 3 else 'Inline sentence addition'
        )
        rows.append((url, kw, ', '.join(sorted(phrases)), title, action))

    rows.sort(key=lambda r: (r[0], r[1]))

    writer = csv.writer(sys.stdout)
    writer.writerow(['url', 'target_keyword', 'related_words_missing',
                     'current_seo_title', 'suggested_action'])
    for r in rows:
        writer.writerow(r)

    # Stderr summary
    n_total = len(rows)
    actions = defaultdict(int)
    for _, _, _, _, a in rows:
        actions[a] += 1
    sys.stderr.write(f'[seo-semantic-gap] {n_total} (url, keyword) gaps\n')
    for a in sorted(actions):
        sys.stderr.write(f'  {actions[a]:3d}  {a}\n')


if __name__ == '__main__':
    main()
