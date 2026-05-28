#!/usr/bin/env python3
"""
Generate the backlinks outreach workspace CSV from a Semrush "Ideas"
export. Sister to scripts/seo-keyword-audit.mjs — same export, different
slice (the "acquire backlinks from these domains" idea family).

Why Python and not Node: parsing Semrush's XLSX directly is a one-line
openpyxl call versus pulling in an XLSX dependency in Node. The
companion seo-keyword-audit script stays in Node because it operates
on a CSV the merchant has already saved out — the merchant only needs
Python installed to refresh the backlinks list (and openpyxl, which
ships with most Python distros + can be `pip install`'d in 5 sec).

Output columns (single sheet, sorted by priority class then mention count):
  priority_class           — A-publisher / A-directory / A-aggregator
                             / B-niche / B-evaluate / C-competitor /
                             C-low-value / D-search / D-self / D-noise
  target_domain            — the suggested backlink source
  mention_count            — how many Semrush ideas reference this
                             domain (rough proxy for relevance)
  our_pages_to_pitch       — pages on our site that should be linked
                             (top 5 per domain by frequency)
  keywords_we_want_to_rank_for
                           — keywords the page is targeting (top 8
                             per domain)
  outreach_notes           — manual classification + outreach angle

Usage:
  python3 scripts/seo-backlinks-outreach.py path/to/semrush-ideas.xlsx \\
    > docs/seo-audits/backlinks-outreach-YYYYMMDD.csv

The CLASSIFICATION dict below is the merchant-curated mapping of
domain → (tier, outreach note). Extend it as new target domains
appear in future exports — anything not pre-classified falls into
`B-evaluate` so it doesn't get silently dropped.
"""

import csv
import re
import sys
from collections import defaultdict, Counter
from openpyxl import load_workbook

CLASSIFICATION = {
    # Tier A — real outreach targets, likely to link, high domain authority
    'hfbusiness.com':                ('A-publisher',  'Home Furnishings Business trade pub. Pitch product news or expert column.'),
    'futureofpersonalhealth.com':    ('A-publisher',  'Mediaplanet health pub. Pitch sleep-health expert quote.'),
    'eatthis.com':                   ('A-publisher',  'Health/lifestyle outlet. Pitch a "best mattress for X" round-up inclusion.'),
    'designrulz.com':                ('A-publisher',  'Design blog covering bedroom + mattress topics. Pitch a guest mattress-buying piece.'),
    'weekand.com':                   ('A-publisher',  'USA Today network lifestyle. Pitch sleep tips or mattress care.'),
    'threebestrated.com':            ('A-directory',  'Local-biz directory. Apply for LA mattress-store category (annual audit).'),
    'wheretobuyamattress.com':       ('A-directory',  'Mattress directory. Submit your 5 LA showrooms.'),
    'redditrecs.com':                ('A-aggregator', 'Reddit aggregator. Strategy: improve Reddit mentions (r/Mattress, r/LosAngeles).'),
    'yahoo.com':                     ('A-aggregator', 'Yahoo News + Shopping syndicate. Pitch press releases to PRWeb/Cision; coverage often appears on Yahoo properties.'),
    'shopmattressondemand.com':      ('A-aggregator', 'Mattress comparison aggregator. Submit your products + ask for inclusion.'),

    # Tier B — niche / lower priority
    'mondoro.com':                   ('B-niche',      'B2B furniture sourcing blog. Less SEO value vs consumer pubs; pitch a manufacturing-quality piece if you have one.'),
    'loveyourjester.com':            ('B-niche',      'Small lifestyle blog. Pitch a sleep-tips guest post.'),
    'milled.com':                    ('B-niche',      'Email-newsletter archive. Listing happens automatically when retailers send promotional emails — no outreach needed.'),
    'deialiving.com':                ('B-niche',      'Lifestyle/home-decor blog. Worth a guest-post pitch.'),

    # Tier C — competitors or low transfer value
    'nationalmattress.ca':           ('C-competitor', 'Canadian mattress competitor. Unlikely to link; geo mismatch.'),
    'biltritefurniture.com':         ('C-competitor', 'NY furniture competitor. Won\'t link.'),
    'jamesandjamesfurniture.com':    ('C-competitor', 'UK furniture competitor. Geo + competition mismatch.'),
    'alibaba.com':                   ('C-low-value',  'Supplier directory. Marginal SEO value.'),
    'sourcifychina.com':             ('C-low-value',  'B2B supplier marketplace; wrong audience.'),

    # Tier D — noise, ignore
    'bing.com':                      ('D-search',     'Search engine, not a backlink path. Skip.'),
    'mattressstoreslosangeles.com':  ('D-self',       'YOUR OWN DOMAIN — Semrush self-reference bug. Skip.'),
    'one-tab.com':                   ('D-noise',      'Browser-extension tab collections. No SEO value.'),
    'accio.com':                     ('D-noise',      'AI shopping assistant. No editorial backlink path.'),
    'keywordseverywhere.com':        ('D-noise',      'SEO tool. Their backlinks are from app installs, not editorial.'),
    'parse.gl':                      ('D-noise',      'URL utility / parser. No SEO value.'),
    'glarity.app':                   ('D-noise',      'AI summarizer app. No editorial backlink path.'),
    'gaminodena.com':                ('D-noise',      'Unclear / personal blog. Skip unless researched.'),
    'hol.es':                        ('D-noise',      'Spanish-language site, geo + language mismatch.'),
    'nuxx.net':                      ('D-noise',      'Personal tech blog. Off-topic; skip.'),
    'emailinspire.com':              ('D-noise',      'Email-template marketplace. Off-topic.'),
    'emailsnest.com':                ('D-noise',      'Email-archive site. Same as milled.com — automatic.'),
    'rioseo.com':                    ('D-noise',      'Local SEO tooling vendor. Off-topic.'),
    'grokipedia.com':                ('D-noise',      'New wiki-style platform. Marginal authority.'),
}

DOMAIN_RE = re.compile(r'^[a-z0-9.-]+\.[a-z]{2,}$')


def main():
    if len(sys.argv) < 2:
        sys.stderr.write('Usage: seo-backlinks-outreach.py <semrush-ideas.xlsx>\n')
        sys.exit(2)

    wb = load_workbook(sys.argv[1], data_only=True)
    # Semrush exports name the sheet "SEO Ideas" — fall back to the
    # first sheet if a future export uses a different name.
    ws = wb['SEO Ideas'] if 'SEO Ideas' in wb.sheetnames else wb[wb.sheetnames[0]]

    by_domain = defaultdict(list)
    site_prefixes = (
        'https://www.mattressstoreslosangeles.com',
        'https://mattressstoreslosangeles.com',
    )
    for r in range(2, ws.max_row + 1):
        idea = ws.cell(row=r, column=4).value or ''
        url = (ws.cell(row=r, column=2).value or '')
        kw = ws.cell(row=r, column=3).value or ''
        for prefix in site_prefixes:
            url = url.replace(prefix, '')
        if 'acquire backlinks' not in idea:
            continue
        # Format: "Try to acquire backlinks from the following domains: a.com, b.com, ..."
        suffix = idea.split('domains:', 1)[-1]
        for raw in suffix.split(','):
            d = raw.strip().rstrip('.').strip().lower()
            if d and DOMAIN_RE.match(d):
                by_domain[d].append((url, kw))

    writer = csv.writer(sys.stdout)
    writer.writerow([
        'priority_class', 'target_domain', 'mention_count',
        'our_pages_to_pitch', 'keywords_we_want_to_rank_for',
        'outreach_notes',
    ])
    rows = []
    for domain, hits in by_domain.items():
        klass, note = CLASSIFICATION.get(
            domain,
            ('B-evaluate', 'Not pre-classified — research site, decide if worth pursuing.'),
        )
        pages = sorted(set(h[0] for h in hits))
        kws = sorted(set(h[1] for h in hits))
        rows.append((klass, domain, len(hits), '; '.join(pages[:5]), '; '.join(kws[:8]), note))
    rows.sort(key=lambda r: (r[0], -r[2]))
    for r in rows:
        writer.writerow(r)

    # Stderr summary so the human running the script sees a recap.
    classes = Counter(CLASSIFICATION.get(d, ('B-evaluate',))[0] for d in by_domain)
    total = sum(len(v) for v in by_domain.values())
    sys.stderr.write(
        f'[seo-backlinks-outreach] {len(by_domain)} unique domains, '
        f'{total} mentions\n'
    )
    for k, v in sorted(classes.items()):
        sys.stderr.write(f'  {k}: {v} domains\n')


if __name__ == '__main__':
    main()
